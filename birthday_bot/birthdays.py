from __future__ import annotations

import calendar
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

LOGGER = logging.getLogger(__name__)

MONTHS = {
    "января": 1,
    "январь": 1,
    "февраля": 2,
    "февраль": 2,
    "марта": 3,
    "март": 3,
    "апреля": 4,
    "апрель": 4,
    "мая": 5,
    "май": 5,
    "июня": 6,
    "июнь": 6,
    "июля": 7,
    "июль": 7,
    "августа": 8,
    "август": 8,
    "сентября": 9,
    "сентябрь": 9,
    "октября": 10,
    "октябрь": 10,
    "ноября": 11,
    "ноябрь": 11,
    "декабря": 12,
    "декабрь": 12,
}

MONTHS_GENITIVE = {
    1: "января",
    2: "февраля",
    3: "марта",
    4: "апреля",
    5: "мая",
    6: "июня",
    7: "июля",
    8: "августа",
    9: "сентября",
    10: "октября",
    11: "ноября",
    12: "декабря",
}

WEEKDAYS = {
    0: "понедельник",
    1: "вторник",
    2: "среда",
    3: "четверг",
    4: "пятница",
    5: "суббота",
    6: "воскресенье",
}

WEEKDAYS_ACCUSATIVE = {
    0: "понедельник",
    1: "вторник",
    2: "среду",
    3: "четверг",
    4: "пятницу",
    5: "субботу",
    6: "воскресенье",
}


@dataclass(frozen=True)
class BirthdayEntry:
    full_name: str
    department: str
    day: int
    month: int
    entry_id: Optional[int] = None


@dataclass(frozen=True)
class ScheduledBirthday:
    target_date: date
    entry: BirthdayEntry


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _parse_int(value: object) -> Optional[int]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)

    text = normalize_text(value)
    if not text:
        return None

    match = re.search(r"\d{1,2}", text)
    return int(match.group()) if match else None


def _parse_month_name(value: object) -> Optional[int]:
    text = normalize_text(value).lower().replace(".", " ")
    if not text:
        return None

    for token in text.split():
        if token in MONTHS:
            return MONTHS[token]
    return None


def is_valid_day_month(day: int, month: int, year: int = 2024) -> bool:
    if day < 1 or month < 1 or month > 12:
        return False
    _, max_day = calendar.monthrange(year, month)
    return day <= max_day


def format_day_month(day: int, month: int) -> str:
    return f"{day} {MONTHS_GENITIVE[month]}"


def parse_day_month_text(value: str) -> tuple[int, int]:
    text = normalize_text(value)
    if not text:
        raise ValueError("Дата не указана.")

    numeric_match = re.fullmatch(r"(\d{1,2})[./-](\d{1,2})(?:[./-]\d{2,4})?", text)
    if numeric_match:
        day = int(numeric_match.group(1))
        month = int(numeric_match.group(2))
        if is_valid_day_month(day, month):
            return day, month
        raise ValueError("Дата должна быть в формате ДД.ММ и быть корректной.")

    day = _parse_int(text)
    month = _parse_month_name(text)
    if day is None or month is None or not is_valid_day_month(day, month):
        raise ValueError("Дата должна быть в формате ДД.ММ или, например, 7 апреля.")
    return day, month


def _parse_date_cell(value: object) -> tuple[Optional[int], Optional[int]]:
    if value is None or value == "":
        return None, None

    if isinstance(value, datetime):
        return value.day, value.month
    if isinstance(value, date):
        return value.day, value.month
    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value)
        except Exception:
            return None, None
        return parsed.day, parsed.month

    text = normalize_text(value)
    if not text:
        return None, None

    try:
        return parse_day_month_text(text)
    except ValueError:
        return None, None


def _extract_day_month(row: tuple[object, ...]) -> tuple[Optional[int], Optional[int]]:
    day = _parse_int(row[3] if len(row) > 3 else None)
    month = _parse_int(row[4] if len(row) > 4 else None)
    if day and month:
        return day, month

    date_cell = row[2] if len(row) > 2 else None
    return _parse_date_cell(date_cell)


def _sorted_entries(entries: Iterable[BirthdayEntry]) -> list[BirthdayEntry]:
    return sorted(
        entries,
        key=lambda item: (item.month, item.day, item.full_name.lower(), item.entry_id or 0),
    )


def _row_to_entry(index: int, row: tuple[object, ...]) -> Optional[BirthdayEntry]:
    full_name = normalize_text(row[0] if len(row) > 0 else None)
    if not full_name:
        return None

    department = normalize_text(row[1] if len(row) > 1 else None)
    raw_date_values = [
        row[2] if len(row) > 2 else None,
        row[3] if len(row) > 3 else None,
        row[4] if len(row) > 4 else None,
    ]
    day, month = _extract_day_month(row)
    if day is None or month is None:
        if any(value not in (None, "") for value in raw_date_values):
            LOGGER.warning("Пропущена строка %s: не удалось разобрать дату для %s", index, full_name)
        return None

    if not is_valid_day_month(day, month):
        LOGGER.warning("Пропущена строка %s: не удалось разобрать дату для %s", index, full_name)
        return None

    return BirthdayEntry(
        full_name=full_name,
        department=department,
        day=day,
        month=month,
        entry_id=index,
    )


def load_birthdays_from_workbook(workbook_path: Path) -> list[BirthdayEntry]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    birthdays: list[BirthdayEntry] = []

    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if index == 1:
            continue

        entry = _row_to_entry(index, row)
        if entry is not None:
            birthdays.append(entry)

    workbook.close()
    return _sorted_entries(birthdays)


def observed_birthday(entry: BirthdayEntry, year: int) -> date:
    if entry.month == 2 and entry.day == 29 and not calendar.isleap(year):
        return date(year, 2, 28)
    return date(year, entry.month, entry.day)


def weekly_birthdays_for_notification(
    entries: Iterable[BirthdayEntry],
    dispatch_date: date,
) -> list[ScheduledBirthday]:
    scheduled: list[ScheduledBirthday] = []
    normalized_entries = list(entries)

    for offset in range(1, 10):
        target_date = dispatch_date + timedelta(days=offset)
        day_entries = [
            ScheduledBirthday(target_date=target_date, entry=entry)
            for entry in normalized_entries
            if observed_birthday(entry, target_date.year) == target_date
        ]
        scheduled.extend(
            sorted(
                day_entries,
                key=lambda item: (item.entry.full_name.lower(), item.entry.entry_id or 0),
            )
        )

    return scheduled


def birthdays_for_date(
    entries: Iterable[BirthdayEntry],
    target_date: date,
) -> list[BirthdayEntry]:
    matches = [
        entry
        for entry in entries
        if observed_birthday(entry, target_date.year) == target_date
    ]
    return _sorted_entries(matches)


def format_target_date(target_date: date) -> str:
    return format_day_month(target_date.day, target_date.month)


def build_weekly_reminder_lines(
    scheduled_birthdays: list[ScheduledBirthday],
    dispatch_date: date,
) -> list[str]:
    if not scheduled_birthdays:
        return ["Дней рождений в Сбере на следующей неделе нет."]

    period_start = dispatch_date + timedelta(days=1)
    period_end = dispatch_date + timedelta(days=9)
    lines = [
        "Напоминание по дням рождения.",
        (
            f"Рассылка за {WEEKDAYS_ACCUSATIVE[dispatch_date.weekday()]}, {format_target_date(dispatch_date)}: "
            f"период с {format_target_date(period_start)} по {format_target_date(period_end)}."
        ),
    ]

    current_date: Optional[date] = None
    for scheduled in scheduled_birthdays:
        if scheduled.target_date != current_date:
            current_date = scheduled.target_date
            lines.extend(
                [
                    "",
                    f"{WEEKDAYS[current_date.weekday()].capitalize()}, {format_target_date(current_date)}:",
                ]
            )

        suffix = f" ({scheduled.entry.department})" if scheduled.entry.department else ""
        lines.append(f"- {scheduled.entry.full_name}{suffix}")

    return lines


def build_today_birthday_lines(entries: list[BirthdayEntry], target_date: date) -> list[str]:
    lines = [f"День рождения сегодня - {format_target_date(target_date)} у:"]
    for entry in entries:
        suffix = f" ({entry.department})" if entry.department else ""
        lines.append(f"- {entry.full_name}{suffix}")
    return lines


def build_birthday_list_lines(entries: list[BirthdayEntry]) -> list[str]:
    lines = [f"Всего дней рождений: {len(entries)}", ""]
    for entry in entries:
        suffix = f" ({entry.department})" if entry.department else ""
        lines.append(
            f"#{entry.entry_id} {format_day_month(entry.day, entry.month)} - "
            f"{entry.full_name}{suffix}"
        )
    return lines
